# -*- coding: utf-8 -*-
# file_name       : excel_write.py
# file_func       :
# file_author     : 'Johnathan.Wick'
# file_create_date: 2020/5/26 20:51

import time
from datetime import datetime

import pandas as pd
from pandas import DataFrame
from openpyxl import Workbook as Owb
from pyexcelerate import Workbook, Color, Style, Font, Fill, Format

length = 1000
src_data = [[i for i in range(12)]] * length + [[i for i in range(12, 24)]]


def time_wrapper(times=1):
    def fuck(fn):
        def inner(*args, **kwargs):
            total = 0
            for _ in range(times):
                s = time.time()
                fn(*args, **kwargs)
                cost = time.time() - s
                print(f'{fn} total cost ------------->', cost)
                total += cost
            print(f'avg ---------->{total / times}')

        return inner

    return fuck


@time_wrapper()
def write_by_py_accelerate(file_path):
    wb = Workbook()
    ws = wb.new_sheet('fucking_test')
    ws.range('A2', 'L%d' % (length + 1)).value = src_data
    # ws.range("A2", "L10001").style.font.color = 'green'
    # ws.range("A2", "L10001").style.font.bold = True
    # ws.range("A2", "L10001").style.font.italic = True
    # for i in range(1, length + 2):
    #     ws.set_row_style(i, Style(font=Font(bold=True, italic=True)))
    # for i in range(1, 13):
    #     ws.set_col_style(i, Style(font=Font(bold=True, italic=True)))
    wb.save(file_path)


acc_file_path = r'e:\acclerate.xlsx'


# write_by_py_accelerate(acc_file_path)


@time_wrapper
def write_by_openpyxl(file_path):
    wb = Owb()
    wb.create_sheet(title='fucking_Test', index=0)
    ws = wb['fucking_Test']
    for item in src_data:
        ws.append(item)
    wb.save(file_path)


openpyxl_path = r'e:\openpyxl.xlsx'


# write_by_openpyxl(openpyxl_path)


@time_wrapper()
def write_by_pandas(file_path):
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter')
    df = DataFrame(src_data)
    workbook = writer.book
    title_fmt = workbook.add_format(
        {'bold': True, 'font_name': '宋体', 'font_color': 'black', 'align': 'center', 'valign': 'vcenter',
         'font_size': 18, 'bg_color': '#9FC3D1'})
    content_fmt = workbook.add_format({'font_name': '宋体', 'font_color': 'black', 'align': 'center',
                                       'font_size': 12})
    link_fmt = workbook.add_format({'font_name': '宋体', 'font_color': 'blue', 'align': 'center',
                                    'font_size': 12})
    border_format = workbook.add_format({'border': 1})
    for i in range(12):
        sheet_name = f'fucking_test{i}'
        df.to_excel(writer, sheet_name, index=False, header=False)
        worksheet = writer.sheets[sheet_name]
        worksheet.set_column('A:L', 25, content_fmt)
        worksheet.set_default_row(20)
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, title_fmt)
        worksheet.merge_range('K1:L3', u'中文测试', title_fmt)
        worksheet.conditional_format('A1:L%d' % len(df.index), {'type': 'no_blanks', 'format': border_format})
        worksheet.autofilter('K1:K1')
        worksheet.autofilter('A1:A1')
        if not i:
            for j in range(1, 12):
                worksheet.write_url('A%d' % (j + 1), f'internal:fucking_test{j}!A1', string=f'fucking_test{j}',
                                    cell_format=link_fmt)
            # print(worksheet.filter_column(0, 'x = 0'))
    writer.save()


pandas_path = r'E:\pands.xlsx'

write_by_pandas(pandas_path)


def set_styles():
    wb = Workbook()
    ws = wb.new_sheet("sheet name")
    ws.set_cell_value(1, 1, 1)
    ws.set_cell_style(1, 1, Style(font=Font(bold=True)))
    ws.set_cell_style(1, 1, Style(font=Font(italic=True)))
    ws.set_cell_style(1, 1, Style(font=Font(underline=True)))
    ws.set_cell_style(1, 1, Style(font=Font(strikethrough=True)))
    ws.set_cell_style(1, 1, Style(fill=Fill(background=Color(255, 0, 0, 0))))
    ws.set_cell_value(1, 2, datetime.now())
    ws.set_cell_style(1, 1, Style(format=Format('mm/dd/yy')))
    wb.save(r"E:\output.xlsx")

# set_styles()
